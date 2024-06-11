from tkinter import *
from tkinter.ttk import Combobox,Treeview
import tkinter as tk
from tkinter import messagebox
import openpyxl , xlrd #Python xlrd is used to retrieve information from a spreadsheet; also, python openpyxl reads and writes information from the spreadsheet.
from openpyxl import Workbook
import pathlib # It gathers the necessary functionality in one place and makes it available through methods and properties on a convenient Path object.
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdfcanvas # for pdf
import os 
from datetime import date # for today date
from PIL import Image,ImageTk #for background image


#for setting the background
def set_background_image(root, image_path):
    # Load the image
    image = Image.open(image_path)
    photo = ImageTk.PhotoImage(image)

    window_width = 1200
    window_height = 600

    # Resize the image to fit the window size
    resized_image = image.resize((window_width, window_height))
    photo = ImageTk.PhotoImage(resized_image)


    # Create a canvas widget
    background = Canvas(root, width=window_width,height=window_height)
    background.pack(fill='both', expand=True)

    # Set the image as the background
    background.create_image(0, 0, image=photo, anchor='nw')

    # Add widgets on the canvas
    background.image = photo  # Keep a reference to the image to prevent garbage collection

    return background



# main window
root=Tk()
root.title("WHTM's Hostel Service")
root.geometry('1200x600+50+50')
root.resizable(False,False) 
root.configure(bg="#FF9966")

background_image_path = "img1.jpeg"  # Replace with your image file path

# Set the background image
canvas = set_background_image(root, background_image_path)


# data file path
backened='backened_data.xlsx'
room_file_path = 'rooms.xlsx'
services_file = "services.xlsx"
selected_services_file = "selected_services.xlsx"
pdf_bill_file = "bill.pdf"
room_data = {}
TODAY=date.today()

#for loading the room data
def load_room_data():
    global room_data
    file = pathlib.Path(room_file_path)
    if file.exists():
        workbook = openpyxl.load_workbook(room_file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            room, bed1, bed2 = row
            room_data[room] = [bed1, bed2]
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Room', 'Bed1', 'Bed2'])
        for i in range(1, 11):
            sheet.append([f'Room {i}', 'Vacant', 'Vacant'])
        workbook.save(room_file_path)
        load_room_data()

#for allocating the room and bed to the guest
def allocate_room():
    for room, beds in room_data.items():
        for i in range(2):
            if beds[i] == 'Vacant':
                room_data[room][i] = 'Occupied'
                return room, i + 1
    return None, None

#to save the guest data in the rooms file
def save_room_data():
    workbook = openpyxl.load_workbook(room_file_path)
    sheet = workbook.active
    for row_num, (room, beds) in enumerate(room_data.items(), start=2):
        sheet.cell(row=row_num, column=1, value=room)
        sheet.cell(row=row_num, column=2, value=beds[0])
        sheet.cell(row=row_num, column=3, value=beds[1])
    workbook.save(room_file_path)


#Function to show the rooms file on the screen
def show_registration():
    backend_win = Toplevel(root)
    backend_win.title("Reservations")
    backend_win.geometry("1200x600+50+50")
    backend_win.resizable(False, False)
    
    # Load data from the backend file
    file_path = 'Backened_data.xlsx'
    if not os.path.exists(file_path):
        messagebox.showerror("File Not Found", "Backend data file does not exist.")
        backend_win.destroy()
        return
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Create Treeview widget
    tree = Treeview(backend_win, columns=('Full Name', 'Phone Number', 'Age', 'Gender', 'Address', 'Room', 'Bed', 'Check In', 'Check Out', 'Service', 'Total Prize'), show='headings')
    tree.heading('Full Name', text='Full Name')
    tree.heading('Phone Number', text='Phone Number')
    tree.heading('Age', text='Age')
    tree.heading('Gender', text='Gender')
    tree.heading('Address', text='Address')
    tree.heading('Room', text='Room')
    tree.heading('Bed', text='Bed')
    tree.heading('Check In', text='Check In')
    tree.heading('Check Out', text='Check Out')
    tree.heading('Service', text='Service')
    tree.heading('Total Prize', text='Total Prize')
    
    # Insert data into Treeview
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)
    
    tree.pack(expand=YES, fill=BOTH)


#ToShow room Data
def show_room():
    room_win = Toplevel(root)
    room_win.title("Rooms Data")
    room_win.geometry("700x500+50+50")
    room_win.resizable(False, False)
    
    # Load data from the backend file
    file_path = 'Rooms.xlsx'
    if not os.path.exists(file_path):
        messagebox.showerror("File Not Found", "Rooms file does not exist.")
        room_win.destroy()
        return
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Create Treeview widget
    tree = Treeview(room_win, columns=('Room', 'Bed1', 'Bed2'), show='headings')
    tree.heading('Room', text='Room')
    tree.heading('Bed1', text='Bed1')
    tree.heading('Bed2', text='Bed2')
    
    # Insert data into Treeview
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)
    
    tree.pack(expand=YES, fill=BOTH)


# for new registration Window
def new_register():
        def submit():
            name=nameValue.get()
            contact=ContactValue.get()
            Cnic=CnicValue.get()
            gender=gender_combobox.get()
            address=addressEntry.get(1.0,END)
            date=DateValue.get()

            if not name or not contact or not CnicValue or not address:
                messagebox.showerror('Invalid Entry', 'All fields are required')
                return
            calculate_total_clicked(checkboxes,name)
            room, bed = allocate_room()
            if not room:
                messagebox.showerror('No Vacancy', 'All rooms are fully occupied')
                return
            
            #Saving the information to the file
            file=openpyxl.load_workbook('Backened_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=name)
            sheet.cell(column=2,row=sheet.max_row,value=contact)
            sheet.cell(column=3,row=sheet.max_row,value=Cnic)
            sheet.cell(column=4,row=sheet.max_row,value=gender)
            sheet.cell(column=5,row=sheet.max_row,value=address)
            sheet.cell(column=6,row=sheet.max_row,value=room)
            sheet.cell(column=7,row=sheet.max_row,value=bed)
            sheet.cell(column=8,row=sheet.max_row,value=date)
            sheet.cell(column=10, row=sheet.max_row, value=len(selected_services))
            sheet.cell(column=11, row=sheet.max_row, value=total)

            file.save(r'Backened_data.xlsx')
            save_room_data()
            messagebox.showinfo('info',f'Detail added! Room: {room}, Bed: {bed}')

            nameValue.set('')
            ContactValue.set('')
            CnicValue.set('')
            DateValue.set('')
            addressEntry.delete(1.0,END)
            

        # To load the services from the services file.   
        def load_services():
                services = {}
                if os.path.exists(services_file):
                    workbook = openpyxl.load_workbook(services_file)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        service, price = row
                        services[service] = price
                return services
        

        # Function to handle the "Calculate Total" button click event
        def calculate_total_clicked(checkboxes,name):
                    global selected_services
                    selected_services = []

                    for service, var in checkboxes.items():
                        if var.get() == 1:
                            selected_services.append(service)

                    calculate_total(selected_services,name)

        # Function to calculate total cost per day based on selected services
        def calculate_total(selected_services,name):
                    services = load_services()
                    selected_services_with_prices = {service: services[service] for service in selected_services}
                    global total
                    total = sum(selected_services_with_prices.values())
                    # save_selected_services_to_excel(selected_services)
                    generate_pdf_bill(selected_services_with_prices, total,name)
                    messagebox.showinfo("Bill Info", f"Bill generated successfully for {name}")
                    # guest_names = load_guest_names()
                    services = load_services()

        # Function to generate a PDF bill
        def generate_pdf_bill(selected_services, total_cost,name):
                    # Set custom page size for billing slip
                    page_width = 200
                    page_height = 500
                    pdf_canvas = pdfcanvas.Canvas(pdf_bill_file, pagesize=(page_width, page_height))

                    text_margin = 5
                    text_x = text_margin
                    text_y = page_height - text_margin
                    pdf_canvas.drawString(20, page_height - 50, "WHTM's Hostel Service Bill")
                    pdf_canvas.line(10,440,190,440)
                    pdf_canvas.drawString(20, page_height - 80, f"Name: {name}" )
                    pdf_canvas.drawString(20, page_height - 100 ,f'Date: {TODAY}')
                    pdf_canvas.line(10,390,190,390)
                    pdf_canvas.drawString(20, page_height - 140, "Selected Services:")
                    
                    y = page_height - 160
                    for service, price in selected_services.items():
                        pdf_canvas.drawString(20, y, f"{service}: ${price}")
                        y -= 20
                    pdf_canvas.drawString(20, y, f"Total Cost per Day: ${total_cost}")
                    pdf_canvas.save()


        def clear():
            nameValue.set('')
            ContactValue.set('')
            CnicValue.set('')
            DateValue.set('')
            addressEntry.delete(1.0,END)
               
        
        #Making up the new window for registration
        register_win=Toplevel(root)
        register_win.title("New Regisration of Guest")
        register_win.geometry('900x500+300+100')
        register_win.resizable(False,False) 
        register_win.configure(bg="#FDD017")

        icon_image=PhotoImage(file="logo.png") 
        register_win.iconphoto(False,icon_image)

        #backend  data allotment
        file=pathlib.Path('Backened_data.xlsx')
        if file.exists():
            pass 
        else:
            file=Workbook( )
            sheet=file.active
            
            sheet['A1']='Full Name'
            sheet['B1']='Phone Number'
            sheet['C1']='Cnic'
            sheet['D1']='Gender'
            sheet['E1']='Address'
            sheet['F1']='Room'
            sheet['G1']= 'Bed'
            sheet['H1']= 'Check In'
            sheet['I1']= 'Check out'
            sheet['J1']= 'Service'
            sheet['K1']= 'Total Prize'
            


            file.save('Backened_data.xlsx')
        #heading
        Label(register_win,text="Please fill out this Entry form:",font="arial 13 bold ", bg="#FDD017", fg="black").place(x=300,y=20)

        #label
        Label(register_win,text="Name:",font=25, bg="#FDD017", fg="black").place(x=300,y=100)
        Label(register_win,text="Contact No. :",font=25, bg="#FDD017", fg="black").place(x=300,y=150)
        Label(register_win,text="CNIC:",font=25, bg="#FDD017", fg="black").place(x=300,y=200)
        Label(register_win,text="Check In Date:",font=25, bg="#FDD017", fg="black").place(x=300,y=250)       
        Label(register_win,text="Gender:",font=25, bg="#FDD017", fg="black").place(x=300,y=300)
        Label(register_win,text="Address:",font=25, bg="#FDD017", fg="black").place(x=300,y=350)

        #Entry
        global nameValue
        nameValue=StringVar()
        ContactValue=StringVar()
        CnicValue=StringVar()
        DateValue =StringVar()

        nameEntry = Entry(register_win,textvariable=nameValue ,width=45,bd=2,font=20)
        contactEntry = Entry(register_win,textvariable=ContactValue ,width=45,bd=2,font=20)
        CnicEntry = Entry(register_win,textvariable=CnicValue ,width=45,bd=2,font=20)
        DateEntry = Entry(register_win,text=TODAY,textvariable=DateValue ,width=43,bd=2,font=20)
        
        #gender
        gender_combobox =Combobox(register_win,values=['Male','Female','Others'],font='arial 14',state='r',width=14 )
        gender_combobox.place(x=400,y=300)
        gender_combobox.set('Male')

        addressEntry=Text(register_win,width=50,height=4,bd=4)

        nameEntry.place(x=400,y=100)
        contactEntry.place(x=400,y=150)
        CnicEntry.place(x=400,y=200)
        DateEntry.place(x=420, y=250)
        addressEntry.place(x=400,y=350)

        services = load_services()
        checkboxes = {}
        row = 0
        for service, price in services.items():
            var = IntVar()
            checkbox = Checkbutton(register_win, text=f"{service} (${price}/day)", variable=var)
            checkbox.grid(row=row, column=0, sticky=W)
            checkbox.config(bg="#FDD017", fg="black", font=("Arial", 12), padx=10, pady=5)
            checkboxes[service] = var
            row += 1

    

        Button(register_win,text='Sumbit',bg='black',fg='white',width=15 ,height=2,command=submit).place(x=300,y=440)
        Button(register_win,text='Clear',bg='black',fg='white',width=15 ,height=2,command=clear ).place(x=420,y=440)
        Button(register_win,text='Exit',bg='black',fg='white',width=15 ,height=2 ,command=lambda:register_win.destroy() ).place(x=540,y=440)


        register_win.mainloop() 

# Function to see the guest details
def retrieve_guest():
    check_info = Toplevel(root)
    check_info.title("Retrieve Guest Information")
    check_info.geometry('500x430+500+100')
    check_info.resizable(False, False)
    check_info.configure(bg="#FDD017")
    
    icon_image=PhotoImage(file="logo.png") 
    check_info.iconphoto(False,icon_image)

    #to show the names in the dropdown box
    def load_guest_names():
        file = openpyxl.load_workbook('Backened_data.xlsx')
        sheet = file.active
        guest_names = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            guest_names.append(row[0])
        return guest_names

    guest_names = load_guest_names()

    #show thw specific guest details
    def show_guest_details():
        selected_name = name_combobox.get()
        if not selected_name:
            messagebox.showerror('Invalid Selection', 'Please select a guest name')
            return

        file = openpyxl.load_workbook('Backened_data.xlsx')
        sheet = file.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == selected_name:
                details = f"Name: {row[0]}\nContact: {row[1]}\nCNIC No.: {row[2]}\nGender: {row[3]}\nAddress: {row[4]}\nRoom: {row[5]}\nBed: {row[6]}\nCheck In:{row[7]}\nCheck Out: {row[8]}\nServices: {row[9]}\nTotal Price: {row[10]}"
                details_text.delete(1.0, END)
                details_text.insert(END, details)
                return

        messagebox.showerror('Not Found', 'Guest not found')
    
    #main window
    Label(check_info, text="Select a guest to retrieve details:", font="arial 13 bold ", bg="#FDD017", fg="black").place(x=20, y=20)
    
    name_combobox = Combobox(check_info, values=guest_names, font='arial 14', state='r', width=30)
    name_combobox.place(x=50, y=60)
    
    details_text = Text(check_info, width=55, height=15, bd=4)
    details_text.place(x=20, y=100)

    Button(check_info, text='Show Details', bg='black', fg='white', width=15, height=2, command=show_guest_details).place(x=200, y=360)
    Button(check_info, text='Exit', bg='black', fg='white', width=15, height=2, command=lambda: check_info.destroy()).place(x=350, y=360)

    check_info.mainloop()

# to make checkout of a guest
def delete_guest_data():
    check_out = Toplevel(root)
    check_out.title("Check Out")
    check_out.geometry('650x200+500+100')
    check_out.resizable(False, False)
    check_out.configure(bg="#FDD017")
    #icon
    icon_image=PhotoImage(file="logo.png") 
    check_out.iconphoto(False,icon_image)

    def load_guest_names():
        file = openpyxl.load_workbook('Backened_data.xlsx')
        sheet = file.active
        guest_names = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            guest_names.append(row[0])
        return guest_names

    guest_names = load_guest_names()

    def checkout_guest():
        selected_name = name_combobox.get()
        if not selected_name:
            messagebox.showerror('Invalid Selection', 'Please select a guest name')
            return
        checkout = checkoutValue.get()
        file = openpyxl.load_workbook('Backened_data.xlsx')
        sheet = file.active
        
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == selected_name:
                row_index = row[0].row
                room = row[6].value  # Adjusted column index for room
                bed = row[7].value   # Adjusted column index for bed
                print(f"Selected Guest: {selected_name}, Room: {room}, Bed: {bed}, Checkout Date: {checkout}")
                sheet.cell(column=9, row=row_index, value=checkout)
                break

        file.save('Backened_data.xlsx')

        # Update room data
        if room in room_data:
            if bed == 1:
                room_data[room][0] = 'Vacant'
            elif bed == 2:
                room_data[room][1] = 'Vacant'
            save_room_data()
            print("File saved successfully")

        messagebox.showinfo('Checked Out', f'Guest {selected_name} checked out successfully')
        name_combobox['values'] = load_guest_names()



    checkoutValue=StringVar()
    Label(check_out, text="Select a guest to check out:", font="arial 13 bold ", bg="#FDD017", fg="black").place(x=20, y=20)
    
    name_combobox = Combobox(check_out, values=guest_names, font='arial 14', state='r', width=30)
    name_combobox.place(x=50, y=60)
    Label(check_out,text='Check Out Date:', font="arial 13 bold ", bg="#FDD017", fg="black").place(x=20,y=100)
    checkoutEntry = Entry(check_out,textvariable=checkoutValue,width=45,bd=2,font=20)
    checkoutEntry.place(x=160,y=100)
    Button(check_out, text='Check Out', bg='black', fg='white', width=15, height=2, command=checkout_guest).place(x=200, y=150)
    Button(check_out, text='Exit', bg='black', fg='white', width=15, height=2, command=lambda: check_out.destroy()).place(x=350, y=150)

    check_out.mainloop()


# MAIN SCREEN WINDOW TO DISPLAY FIRST  
#icon
icon_image=PhotoImage(file="logo.png") 
root.iconphoto(False,icon_image)

Label(root, text="Welcome to WHTM's Hostel Service", font=('Arial', 25, 'bold'), bg="white", fg="black").pack(pady=20)

Button(root,text='New Registration', font=23  ,bg='black',fg='white',width=18 ,height=3 ,command=new_register).place(x=40,y=500)
Button(root,text='Rooms', font=23  ,bg='black',fg='white',width=18 ,height=3 , command=show_room).place(x=230,y=500)
Button(root,text='Reservations', font=23  ,bg='black',fg='white',width=18 ,height=3 ,command=show_registration).place(x=420,y=500)
Button(root,text="Guest's Info", font=23  ,bg='black',fg='white',width=18 ,height=3 ,command=retrieve_guest).place(x=610,y=500)
Button(root, text="Check Out",font=23  ,bg='black',fg='white',width=18 ,height=3, command=delete_guest_data).place(x=800,y=500)
Button(root,text='Exit', font=23  ,bg='black',fg='white',width=18 ,height=3 , command=lambda:root.destroy()).place(x=990,y=500)


load_room_data()

root.mainloop() 